"""
Zip-omit attachment splitting.

Broker order emails carry zip-omit lists (xlsx/xls/csv/txt) alongside the order PDF.
Downstream processing needs those zips in batches of at most ZIP_CHUNK_SIZE containing
zip codes and nothing else, so split_zip_file() writes zip-only .xlsx chunks that are
attached alongside the original file (the original is still attached untouched).

Contract mirrors tools_pdf.split_pdf_into_pages: returns (tmp_dir, [paths]) and the
caller deletes tmp_dir when the files have been attached.
"""

import csv
import logging
import os
import re
import tempfile
from pathlib import Path

log = logging.getLogger(__name__)

ZIP_CHUNK_SIZE = 9500

# Share of non-blank cells that must look like zips before a column is treated as
# the zip column (used only when no header names one).
_ZIP_COLUMN_RATIO = 0.9

# Warn when a file explodes into more attachments than this — visible in the log
# rather than quietly adding dozens of files to a ticket.
_CHUNK_WARN_COUNT = 20

_HEADER_RE = re.compile(r"\b(zip|zipcode|postal)", re.IGNORECASE)
_ZIP_RE    = re.compile(r"^(\d{3}|\d{4}|\d{5})(?:-\d{4})?$")
_FIELD_SEP = re.compile(r"[,\t;|]")
_BAD_NAME  = re.compile(r'[\\/:*?"<>|]')


def _zip_digits(value) -> str:
    """Return the bare digits of a zip-shaped value ("332", "1234", "12309"), else ""."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    if not s:
        return ""
    m = _ZIP_RE.match(s)
    return m.group(1) if m else ""


def _pad_digits(digits: str, column_has_five: bool) -> str:
    """
    Turn bare digits into the value to emit.

    Excel stores zips as numbers, so leading zeros are gone by the time we see them:
    4 digits is always a zip that lost one (1234 -> 01234). 3 digits is ambiguous —
    SCF 332 (Orlando) or zip 00332 — so it is resolved from the rest of the column:
    alongside 5-digit zips it is a stripped zip (501 -> 00501, Holtsville), but in a
    column of nothing but 3-digit values it is an SCF list and stays as-is.
    """
    if len(digits) == 5:
        return digits
    if len(digits) == 4:
        return digits.zfill(5)
    return digits.zfill(5) if column_has_five else digits


def _read_rows(path: str) -> list:
    """Read a zip-omit file into a list of row tuples. [] if it cannot be read."""
    suffix = Path(path).suffix.lower()
    try:
        if suffix == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            rows = [r for ws in wb.worksheets for r in ws.iter_rows(values_only=True)]
            wb.close()
            return rows
        if suffix == ".xls":
            import xlrd
            book = xlrd.open_workbook(path)
            return [tuple(sheet.row_values(i))
                    for sheet in book.sheets()
                    for i in range(sheet.nrows)]
        if suffix in (".csv", ".txt"):
            with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
                return [tuple(_FIELD_SEP.split(line.strip()))
                        for line in fh if line.strip()]
    except Exception as e:
        log.warning("Could not read zip file %s: %s", Path(path).name, e)
    return []


def _pick_zip_column(rows: list) -> tuple[int, int]:
    """
    Return (column_index, first_data_row) for the column holding zips.
    (-1, 0) when no column qualifies — i.e. this is not a zip list.

    A header naming the column wins, which is what stops a neighbouring numeric
    column (e.g. the CID column in a 'Zip, Chapter, CID' sheet) from being taken
    for zips. Otherwise the first mostly-zip-shaped column is used.
    """
    if not rows:
        return -1, 0

    def _column_hit_ratio(idx: int, body: list) -> float:
        vals = [r[idx] for r in body if idx < len(r) and str(r[idx] or "").strip()]
        if not vals:
            return 0.0
        return sum(1 for v in vals if _zip_digits(v)) / len(vals)

    header = rows[0]
    for idx, cell in enumerate(header):
        if cell is not None and _HEADER_RE.search(str(cell)):
            # Trust the header only if the column under it actually holds zips.
            if _column_hit_ratio(idx, rows[1:]) >= _ZIP_COLUMN_RATIO:
                return idx, 1

    width = max(len(r) for r in rows)
    for idx in range(width):
        if _column_hit_ratio(idx, rows) >= _ZIP_COLUMN_RATIO:
            return idx, 0

    return -1, 0


def extract_zip_codes(path: str) -> list:
    """
    Zip codes from a .xlsx/.xls/.csv/.txt file, in source order, duplicates kept.
    Returns [] when the file holds no zip column (so non-zip attachments are left alone).
    """
    rows = _read_rows(path)
    if not rows:
        return []

    idx, start = _pick_zip_column(rows)
    if idx < 0:
        log.info("No zip column found in %s — treating as a non-zip attachment",
                 Path(path).name)
        return []

    digits = [_zip_digits(row[idx]) for row in rows[start:] if idx < len(row)]
    digits = [d for d in digits if d]
    # Decide once for the whole column whether 3-digit values are SCFs or zips that
    # lost their leading zeros — see _pad_digits.
    column_has_five = any(len(d) == 5 for d in digits)
    return [_pad_digits(d, column_has_five) for d in digits]


def _write_zip_xlsx(zips: list, dest_path: str) -> None:
    """Write one zip per row in column A, as text so leading zeros survive."""
    import openpyxl
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Zips")
    for z in zips:
        ws.append([z])
    wb.save(dest_path)


def split_zip_file(path: str, display_name: str = "",
                   chunk_size: int = ZIP_CHUNK_SIZE) -> tuple:
    """
    Split a zip-omit file into zip-only .xlsx chunks of at most chunk_size zips.

    Returns (tmp_dir, [chunk_paths]) — caller must delete tmp_dir when done — or
    (None, []) when the file holds chunk_size or fewer zips, or is not a zip list.

    display_name is the original attachment filename; it names the chunks, since the
    file on disk is usually an unhelpful NamedTemporaryFile path.
    """
    zips = extract_zip_codes(path)
    if len(zips) <= chunk_size:
        return None, []

    stem = _BAD_NAME.sub("_", Path(display_name or path).stem)
    total = (len(zips) + chunk_size - 1) // chunk_size
    tmp_dir = tempfile.mkdtemp(prefix="dslf_zips_")
    chunk_paths = []
    for i in range(total):
        chunk = zips[i * chunk_size:(i + 1) * chunk_size]
        dest = os.path.join(tmp_dir, f"{stem}_zips_{i + 1}of{total}.xlsx")
        _write_zip_xlsx(chunk, dest)
        chunk_paths.append(dest)

    log.info("Zip split: %s has %d zips -> %d file(s) of up to %d",
             display_name or Path(path).name, len(zips), total, chunk_size)
    if total > _CHUNK_WARN_COUNT:
        log.warning("Zip split produced %d files from %s — unusually large",
                    total, display_name or Path(path).name)
    return tmp_dir, chunk_paths


def attach_zip_splits(ticket_keys, path: str, display_name: str = "") -> list:
    """
    Split a zip-omit file and attach the chunks to one or more tickets.
    ticket_keys accepts a single key or a list. Returns the chunk filenames attached.

    Never raises — like every other attach step in the pipeline, a failure here is
    logged and the ticket stands.
    """
    import shutil

    if isinstance(ticket_keys, str):
        ticket_keys = [ticket_keys]
    if not ticket_keys:
        return []

    tmp_dir, chunk_paths = None, []
    attached = []
    try:
        tmp_dir, chunk_paths = split_zip_file(path, display_name=display_name)
        if not chunk_paths:
            return []
        from tools_jira import attach_file_to_ticket
        for chunk in chunk_paths:
            name = os.path.basename(chunk)
            for ticket_key in ticket_keys:
                try:
                    attach_file_to_ticket(ticket_key, chunk)
                    log.info("Zip split file attached to %s: %s", ticket_key, name)
                except Exception as e:
                    log.warning("Could not attach zip split %s to %s: %s", name, ticket_key, e)
            attached.append(name)
    except Exception as e:
        log.warning("Zip split failed for %s: %s", display_name or path, e)
    finally:
        if tmp_dir:
            shutil.rmtree(tmp_dir, ignore_errors=True)
    return attached

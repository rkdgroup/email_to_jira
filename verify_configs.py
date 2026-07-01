"""
Audit generated config YAMLs against their source-of-truth files.

Three independent audits:
  1. Excel lookup YAMLs  (full_client_list.yaml + 15 broker YAMLs)
        vs  NEW LR CLIENT LIST 2026.xlsx
  2. client_profiles.yaml
        vs  re-extraction of Client Profiles/**/*.doc(x) (build_profile_yaml.py)
  3. adstra_omit_database.yaml
        vs  Client Profiles/ADSTRA/Adstra Sweeps Client Profile.xlsx

Run from project root:
    python verify_configs.py

Writes config_audit_report.md and prints a summary. Exits non-zero if any
MISSING entry or FIELD MISMATCH is found (EXTRA entries and intentionally
dropped source columns are warnings only).
"""

import re
import sys
import logging
from pathlib import Path

import yaml
import openpyxl

import build_profile_yaml as bpy
from client_lookup import _MANAGER_TO_FILE

logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

_ROOT = Path(__file__).parent
_CONFIG = _ROOT / "config"
_LR_XLSX = _ROOT / "NEW LR CLIENT LIST 2026.xlsx"
_SWEEPS_XLSX = _ROOT / "Client Profiles" / "ADSTRA" / "Adstra Sweeps Client Profile.xlsx"
_REPORT = _ROOT / "config_audit_report.md"

# Excel header (normalized, stripped, upper) -> YAML field
_COL_MAP = {
    "DB": "db_code",
    "DB #": "db_code",
    "BILLING CUST # (PEPBOOK)": "billing_cust",
    "BILLING CUST #": "billing_cust",
    "BILLING CUST # PEPBOOK": "billing_cust",
    "DATABASE NAME": "db_name",
    "RENTAL/EXCHANGE NAME": "rental_name",
    "RENTAL / EXCHANGE NAME": "rental_name",
    "LIST MANAGER": "list_manager",
    "LM CONTACT": "lm_contact",
    "LRR": "lrr",
    "COMMENTS": "comments",
}

# Load-bearing fields (gate the build) vs informational-only fields.
_CORE_FIELDS = ("billing_cust", "db_name", "rental_name", "list_manager")
_INFO_FIELDS = ("lm_contact", "lrr", "comments")

# Known source-sheet data-quality issues where the YAML is deliberately correct.
# Suppressed from the gate (still shown, tagged ACCEPTED). (yaml_base, db_code, field);
# field "__missing__" = a source db_code intentionally absent from the YAML.
_ACCEPTED = {
    ("full_client_list", "C12",  "__missing__"),   # source dropped D suffix; yaml has C12D
    ("full_client_list", "S52D", "billing_cust"),  # source 'S52D' -> billable base 'S52'
    ("rmi",              "J44D", "billing_cust"),   # source 'J44D' -> billable base 'J44'
    # Mature Health: dedicated WASHINGTON LIST tab says M24R (authoritative, and what
    # the runtime uses); the FULL master roll-up is stale at M24. YAMLs use M24R.
    ("full_client_list", "M24N", "billing_cust"),
    ("full_client_list", "M24O", "billing_cust"),
    ("full_client_list", "M24R", "billing_cust"),
}

# Every broker YAML has a dedicated tab in the workbook. Map basename -> sheet.
_YAML_TO_SHEET = {
    "full_client_list": "LIST RENTAL FULL CLIENT SHEET",
    "adstra":           "ADSTRA",
    "aalc":             "AALC",
    "amlc":             "AMLC",
    "celco":            "CELCO",
    "conrad":           "CONRAD",
    "data_axle":        "DATA-AXLE",
    "kap":              "KAP",
    "mary_e_granger":   "MARY E GRANGER",
    "negev":            "NEGEV",
    "nitn":             "NITN",
    "rkd":              "RKD",
    "rmi":              "RMI",
    "washington_list":  "WASHINGTON LIST",
    "we_are_moore":     "WE ARE MOORE",
}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _norm(v) -> str:
    """Normalize a cell/field value for comparison: None/empty -> '', collapse ws."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return re.sub(r"\s+", " ", s)


def _load_sheet(wb, sheet_name: str) -> list[dict]:
    """Return list of {yaml_field: value} dicts for a sheet, mapped by header name."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [(_norm(h).upper() if h is not None else None) for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c is not None and _norm(c) for c in r):
            continue
        rec = {}
        for h, val in zip(headers, r):
            if h is None:
                continue
            field = _COL_MAP.get(h)
            if field:
                rec[field] = _norm(val)
        if rec.get("db_code"):
            out.append(rec)
    return out


def _load_yaml_list(path: Path) -> list[dict]:
    if not path.exists():
        return []
    for enc in ("utf-8", "cp1252"):
        try:
            with open(path, encoding=enc) as f:
                return yaml.safe_load(f) or []
        except UnicodeDecodeError:
            continue
    return []


def _group_by_code(rows: list[dict]) -> dict[str, list[dict]]:
    """db_code (upper) -> list of rows (source has repeated codes: T11R x32, etc.)."""
    from collections import defaultdict
    idx: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        code = _norm(r.get("db_code")).upper()
        if code:
            idx[code].append(r)
    return idx


def _strip_paren(s: str) -> str:
    """'T11 (A42D)' -> 'T11'  (source billing cust sometimes carries a note)."""
    return re.sub(r"\s*\([^)]*\)\s*", " ", s).strip()


def _name_key(v) -> str:
    """Normalize a name for matching: drop a trailing '(ACR)' acronym, lowercase."""
    s = _norm(v)
    s = re.sub(r"\s*\([^)]*\)\s*$", "", s)
    return s.lower()


# ---------------------------------------------------------------------------
# Module 1 — Excel lookup YAMLs
# ---------------------------------------------------------------------------

def audit_excel_lookup() -> dict:
    """Compare every Excel-derived lookup YAML against its dedicated source tab."""
    wb = openpyxl.load_workbook(_LR_XLSX, data_only=True, read_only=True)
    sheet_names = {s.upper(): s for s in wb.sheetnames}
    results = {}
    for yaml_base, sheet_upper in _YAML_TO_SHEET.items():
        actual = sheet_names.get(sheet_upper.upper())
        if not actual:
            results[yaml_base] = {"error": f"source sheet '{sheet_upper}' not found"}
            continue
        src = _load_sheet(wb, actual)
        ycfg = _load_yaml_list(_CONFIG / f"{yaml_base}.yaml")
        results[yaml_base] = _compare_rows(src, ycfg, source_label=f"sheet '{actual}'",
                                           base=yaml_base)
    wb.close()
    return results


def _compare_rows(src_rows: list[dict], yaml_rows: list[dict],
                  source_label: str, base: str = "") -> dict:
    """Collision-aware compare: db_code repeats in source, so a YAML value is a
    match if it equals the value of ANY source row sharing that db_code."""
    src_g = _group_by_code(src_rows)
    yaml_g = _group_by_code(yaml_rows)

    missing_all = sorted(set(src_g) - set(yaml_g))  # source code absent from yaml
    extra = sorted(set(yaml_g) - set(src_g))        # yaml code absent from source
    core, info, accepted = [], [], []

    missing = []
    for c in missing_all:
        if (base, c, "__missing__") in _ACCEPTED:
            accepted.append((c, "__missing__"))
        else:
            missing.append(c)

    for code in sorted(set(src_g) & set(yaml_g)):
        srows = src_g[code]
        for y in yaml_g[code]:
            # billing_cust — strip parenthetical notes ('T11 (A42D)' -> 'T11')
            src_bill = {_strip_paren(_norm(r.get("billing_cust"))).upper()
                        for r in srows if _norm(r.get("billing_cust"))}
            yb = _strip_paren(_norm(y.get("billing_cust"))).upper()
            if src_bill and yb and yb not in src_bill:
                if (base, code, "billing_cust") in _ACCEPTED:
                    accepted.append((code, "billing_cust"))
                else:
                    core.append((code, "billing_cust", sorted(src_bill), yb))

            # list_manager
            src_lm = {_norm(r.get("list_manager")).upper()
                      for r in srows if _norm(r.get("list_manager"))}
            ylm = _norm(y.get("list_manager")).upper()
            if src_lm and ylm and ylm not in src_lm:
                core.append((code, "list_manager", sorted(src_lm), ylm))

            # names — only used for fuzzy (word-overlap) matching, so a mismatch
            # degrades match quality but never corrupts a resolved row. Report as
            # informational (non-gating). Match against any source row for the code.
            for f in ("db_name", "rental_name"):
                src_names = {_name_key(r.get(f)) for r in srows if _norm(r.get(f))}
                yv = _name_key(y.get(f))
                if src_names and yv and yv not in src_names:
                    info.append((code, f,
                                 [_norm(r.get(f)) for r in srows][:3], _norm(y.get(f))))

            # informational (unused at runtime): lm_contact, lrr, comments
            for f in _INFO_FIELDS:
                src_vals = {_norm(r.get(f)).upper() for r in srows if _norm(r.get(f))}
                yv = _norm(y.get(f)).upper()
                if src_vals and yv not in src_vals:
                    info.append((code, f, sorted(src_vals), _norm(y.get(f))))

    return {
        "source_label": source_label,
        "src_count": len(src_g),
        "yaml_count": len(yaml_g),
        "missing": [(c, src_g[c][0]) for c in missing],
        "extra": [(c, yaml_g[c][0]) for c in extra],
        "core": core,
        "info": info,
        "accepted": accepted,
    }


# ---------------------------------------------------------------------------
# Module 2 — client_profiles.yaml vs re-extraction of .doc(x)
# ---------------------------------------------------------------------------

def audit_client_profiles() -> dict:
    """Re-extract profile docs in-memory and diff against committed YAML."""
    files = [
        f for f in bpy._PROFILES.rglob("*")
        if f.is_file()
        and not f.name.startswith("~")
        and f.suffix.lower() in (".doc", ".docx")
    ]

    regenerated: dict[str, dict] = {}
    skipped_no_code, skipped_empty, read_errors = [], [], []

    for path in sorted(files):
        stem = path.stem.upper()
        db_code = bpy._extract_db_code(stem)
        if not db_code:
            skipped_no_code.append(path.name)
            continue
        if db_code in regenerated:
            continue  # first file wins (same rule as build())
        try:
            data = (bpy._extract_docx(path) if path.suffix.lower() == ".docx"
                    else bpy._extract_doc(path))
        except Exception as exc:
            read_errors.append(f"{path.name}: {exc}")
            continue
        if not any([data["select_by"], data["flags"], data["dollar_cap"]]):
            skipped_empty.append(path.name)
            continue
        regenerated[db_code] = data

    committed = _load_yaml_dict(_CONFIG / "client_profiles.yaml")

    only_regen = sorted(set(regenerated) - set(committed))
    only_committed = sorted(set(committed) - set(regenerated))
    field_diffs = []
    for code in sorted(set(regenerated) & set(committed)):
        r, c = regenerated[code], committed[code]
        diffs = []
        for f in ("select_by", "flags", "dollar_cap",
                  "standard_suppressions", "special_instructions"):
            rv, cv = _norm_profile_field(r.get(f)), _norm_profile_field(c.get(f))
            if rv != cv:
                diffs.append((f, rv, cv))
        if diffs:
            field_diffs.append((code, diffs))

    return {
        "regen_count": len(regenerated),
        "committed_count": len(committed),
        "only_regen": only_regen,
        "only_committed": only_committed,
        "field_diffs": field_diffs,
        "skipped_no_code": skipped_no_code,
        "skipped_empty": skipped_empty,
        "read_errors": read_errors,
        "total_files": len(files),
    }


def _norm_profile_field(v) -> str:
    if isinstance(v, list):
        return " | ".join(_norm(x) for x in v)
    return _norm(v)


def _load_yaml_dict(path: Path) -> dict:
    if not path.exists():
        return {}
    for enc in ("utf-8", "cp1252"):
        try:
            with open(path, encoding=enc) as f:
                data = yaml.safe_load(f) or {}
            return {str(k).upper(): v for k, v in data.items() if isinstance(v, dict)}
        except UnicodeDecodeError:
            continue
    return {}


# ---------------------------------------------------------------------------
# Module 3 — adstra_omit_database.yaml vs Adstra Sweeps xlsx
# ---------------------------------------------------------------------------

def audit_adstra_omit() -> dict:
    wb = openpyxl.load_workbook(_SWEEPS_XLSX, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [(_norm(h) if h is not None else None) for h in rows[0]]

    src = {}            # seed_db -> set(flags)
    src_raw = {}        # seed_db -> raw flag string
    dropped_cols = [h for h in headers
                    if h and h not in ("Aegis Acronym", "Client Code",
                                       "Seed Database", "Adstra List Name", "Flags")]
    for r in rows[1:]:
        d = dict(zip(headers, r))
        seed = _norm(d.get("Seed Database")).upper()
        if not seed:
            continue
        raw = _norm(d.get("Flags"))
        toks = {t for t in re.split(r"[,\s]+", raw) if t}
        src[seed] = toks
        src_raw[seed] = raw
    wb.close()

    omit = _load_yaml_list(_CONFIG / "adstra_omit_database.yaml")
    if isinstance(omit, dict):
        omit = omit.get("adstra_database", [])
    yaml_flags = {}
    for e in omit:
        seed = _norm(e.get("seed_database")).upper()
        if seed and seed not in yaml_flags:
            yaml_flags[seed] = {ch for fl in e.get("flags", []) for ch in str(fl)}

    missing = sorted(set(src) - set(yaml_flags))
    extra = sorted(set(yaml_flags) - set(src))
    flag_diffs = []
    for seed in sorted(set(src) & set(yaml_flags)):
        # Expand any combined source tokens (e.g. "!$") into single chars so
        # the intentional split in the YAML is treated as a match.
        src_chars = {ch for tok in src[seed] for ch in tok}
        if src_chars != yaml_flags[seed]:
            flag_diffs.append((seed, sorted(src_chars), sorted(yaml_flags[seed]),
                               src_raw.get(seed, "")))

    return {
        "src_count": len(src),
        "yaml_count": len(yaml_flags),
        "missing": missing,
        "extra": extra,
        "flag_diffs": flag_diffs,
        "dropped_cols": dropped_cols,
    }


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def _md_excel(results: dict) -> tuple[list[str], int, int, int]:
    lines = ["## 1. Excel lookup YAMLs vs NEW LR CLIENT LIST 2026.xlsx\n"]
    lines.append("| YAML | src | yaml | missing | extra | core diff | info diff |")
    lines.append("|------|----:|-----:|--------:|------:|----------:|----------:|")
    tot_missing = tot_core = tot_extra = 0
    detail = []
    for base in sorted(results):
        r = results[base]
        if "error" in r:
            lines.append(f"| {base} | — | — | — | — | ERROR: {r['error']} | |")
            continue
        nm, nx, nc, ni = (len(r["missing"]), len(r["extra"]),
                          len(r["core"]), len(r["info"]))
        tot_missing += nm; tot_extra += nx; tot_core += nc
        lines.append(f"| {base} | {r['src_count']} | {r['yaml_count']} | "
                     f"{nm} | {nx} | {nc} | {ni} |")
        if nm or nx or nc or ni:
            detail.append(f"\n### {base}.yaml  ({r['source_label']})")
            for code, row in r["missing"]:
                detail.append(f"- **MISSING** `{code}` — "
                              f"{row.get('rental_name','')} / {row.get('db_name','')} "
                              f"(billing={row.get('billing_cust','')}, "
                              f"lm={row.get('list_manager','')})")
            for code, row in r["extra"]:
                detail.append(f"- **EXTRA** `{code}` — {row.get('rental_name','')} "
                              f"/ {row.get('db_name','')} (not in source sheet)")
            for code, f, sv, yv in r["core"]:
                detail.append(f"- **CORE** `{code}`.{f}: source={sv}  yaml=`{yv}`")
            for code, f in r.get("accepted", []):
                detail.append(f"- _ACCEPTED_ `{code}`.{f}: known source typo, YAML correct")
            for code, f, sv, yv in r["info"]:
                detail.append(f"- _info_ `{code}`.{f}: source={sv}  yaml=`{yv}`")
    return lines + detail, tot_missing, tot_core, tot_extra


def _md_profiles(r: dict) -> tuple[list[str], int]:
    lines = ["\n## 2. client_profiles.yaml vs re-extracted .doc(x) profiles\n"]
    lines.append(f"- profile files scanned: **{r['total_files']}**")
    lines.append(f"- regenerated entries: **{r['regen_count']}**, "
                 f"committed entries: **{r['committed_count']}**")
    lines.append(f"- skipped (no db_code in filename): {len(r['skipped_no_code'])}")
    lines.append(f"- skipped (no extractable fields): {len(r['skipped_empty'])}")
    lines.append(f"- read errors: {len(r['read_errors'])}")
    # only_regen = genuinely missing from YAML (real gap). field_diffs are almost all
    # cases where the committed YAML is a curated superset of fresh extraction.
    n_problem = len(r["only_regen"])
    if r["only_regen"]:
        lines.append(f"\n**In docs but MISSING from YAML** ({len(r['only_regen'])}): "
                     + ", ".join(f"`{c}`" for c in r["only_regen"]))
    if r["only_committed"]:
        lines.append(f"\n**In YAML but not re-extracted** "
                     f"({len(r['only_committed'])}, manual entries / OLD-profile dedupe): "
                     + ", ".join(f"`{c}`" for c in r["only_committed"]))
    if r["field_diffs"]:
        lines.append(f"\n**Field differences** ({len(r['field_diffs'])} codes — "
                     f"informational; committed YAML is hand-curated/enriched, do NOT "
                     f"blindly regenerate):")
        for code, diffs in r["field_diffs"][:60]:
            for f, rv, cv in diffs:
                lines.append(f"- `{code}`.{f}: docs=`{rv[:80]}`  yaml=`{cv[:80]}`")
        if len(r["field_diffs"]) > 60:
            lines.append(f"- … and {len(r['field_diffs']) - 60} more")
    if r["read_errors"]:
        lines.append("\n<details><summary>read errors</summary>\n")
        lines += [f"- {e}" for e in r["read_errors"][:40]]
        lines.append("</details>")
    return lines, n_problem


def _md_omit(r: dict) -> tuple[list[str], int]:
    lines = ["\n## 3. adstra_omit_database.yaml vs Adstra Sweeps xlsx\n"]
    lines.append(f"- source seed DBs: **{r['src_count']}**, "
                 f"yaml seed DBs: **{r['yaml_count']}**")
    real = 0
    if r["missing"]:
        real += len(r["missing"])
        lines.append("\n**MISSING (in source, not yaml):** "
                     + ", ".join(f"`{s}`" for s in r["missing"]))
    if r["extra"]:
        lines.append("\n**EXTRA (in yaml, not source):** "
                     + ", ".join(f"`{s}`" for s in r["extra"]))
    if r["flag_diffs"]:
        lines.append("\n**Flag differences** (after expanding combined tokens):")
        for seed, sc, yc, raw in r["flag_diffs"]:
            real += 1
            lines.append(f"- `{seed}`: source=`{raw}` → {sc}  yaml={yc}")
    else:
        lines.append("\n_Flags match (combined `!$` tokens treated as split `!`,`$` — OK)._")
    if r["dropped_cols"]:
        lines.append("\n**Source columns intentionally NOT carried into YAML:** "
                     + ", ".join(f"`{c}`" for c in r["dropped_cols"]))
    return lines, real


def main():
    excel = audit_excel_lookup()
    profiles = audit_client_profiles()
    omit = audit_adstra_omit()

    excel_md, e_missing, e_core, e_extra = _md_excel(excel)
    prof_md, p_problem = _md_profiles(profiles)
    omit_md, o_problem = _md_omit(omit)

    hard = e_missing + e_core + o_problem  # MISSING + core-field diffs gate the build
    header = [
        "# Config audit report",
        "",
        f"- Excel lookup: **{e_missing} missing, {e_core} core-field diffs, "
        f"{e_extra} extra (carry-overs)**",
        f"- client_profiles: **{p_problem} missing from YAML** "
        f"(field diffs are curated-superset, informational)",
        f"- adstra_omit: **{o_problem} real problems**",
        "",
        f"**Gate (MISSING + core diffs): {hard}** "
        + ("[CLEAN]" if hard == 0 else "[NEEDS FIXES]"),
        "",
    ]
    report = "\n".join(header + excel_md + prof_md + omit_md) + "\n"
    _REPORT.write_text(report, encoding="utf-8")

    print("\n".join(header))
    print(f"Full report written to {_REPORT}")
    sys.exit(1 if hard else 0)


if __name__ == "__main__":
    main()
